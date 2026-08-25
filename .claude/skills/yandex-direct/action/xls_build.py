#!/usr/bin/env python3
"""Сборка XLSX для загрузки в Директ из плана кампании.

    python3 xls_build.py campaign.json --out campaign.xlsx

План — JSON вида {"campaigns": [{"name", "type", "groups": [{"name", "keywords",
"negative_keywords", "ads": [{"titles", "texts", "href", "display_path", "sitelinks",
"callouts", "image"}]}]}]}.

Формат листа «Тексты» — docs/alternative-interfaces/xls-interface.md. Пустые ячейки
в повторяющихся колонках означают «то же, что строкой выше»: так Директ понимает,
что строки относятся к одной группе.
"""
import argparse, json, sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("нужен openpyxl: pip install openpyxl")

COLUMNS = [
    ("Название кампании", 26), ("Тип кампании", 18), ("Название группы", 24),
    ("Номер группы", 12), ("Фраза (с минус-словами)", 40), ("Ссылка", 46),
    ("Заголовок 1", 34), ("Заголовок 2", 30), ("Заголовок 3", 30), ("Заголовок 4", 30),
    ("Заголовок 5", 30), ("Текст 1", 44), ("Текст 2", 44), ("Текст 3", 44),
    ("Отображаемая ссылка", 20), ("Регионы", 16), ("Минус-фразы на группу", 40),
    ("Заголовок быстрой ссылки 1", 24), ("Адрес быстрой ссылки 1", 34),
    ("Заголовок быстрой ссылки 2", 24), ("Адрес быстрой ссылки 2", 34),
    ("Заголовок быстрой ссылки 3", 24), ("Адрес быстрой ссылки 3", 34),
    ("Заголовок быстрой ссылки 4", 24), ("Адрес быстрой ссылки 4", 34),
    ("Уточнения", 40), ("ID изображения", 18),
]


def rows_for_group(camp, group, gnum):
    """Строки группы: первая несёт объявление, дальше — только фразы."""
    ads = group.get("ads", [])
    keywords = group.get("keywords", [])
    rows = []
    for i, ad in enumerate(ads):
        titles = (ad.get("titles", []) + [""] * 5)[:5]
        texts = (ad.get("texts", []) + [""] * 3)[:3]
        sl = (ad.get("sitelinks", []) + [{}] * 4)[:4]
        row = {
            "Название кампании": camp["name"] if i == 0 and gnum == 1 else "",
            "Тип кампании": camp.get("type", "Текстово-графические объявления") if i == 0 and gnum == 1 else "",
            "Название группы": group["name"] if i == 0 else "",
            "Номер группы": gnum if i == 0 else "",
            "Фраза (с минус-словами)": keywords[0] if i == 0 and keywords else "",
            "Ссылка": ad.get("href", ""),
            "Отображаемая ссылка": ad.get("display_path", ""),
            "Регионы": camp.get("geo", "") if i == 0 and gnum == 1 else "",
            "Минус-фразы на группу": ", ".join(group.get("negative_keywords", [])) if i == 0 else "",
            "Уточнения": ", ".join(ad.get("callouts", [])),
            "ID изображения": ad.get("image", ""),
        }
        for n, t in enumerate(titles, 1):
            row[f"Заголовок {n}"] = t
        for n, t in enumerate(texts, 1):
            row[f"Текст {n}"] = t
        for n, s in enumerate(sl, 1):
            row[f"Заголовок быстрой ссылки {n}"] = s.get("title", "")
            row[f"Адрес быстрой ссылки {n}"] = s.get("href", "")
        rows.append(row)
    # остальные фразы группы — отдельными строками без объявления
    for kw in keywords[1:]:
        rows.append({"Фраза (с минус-словами)": kw})
    return rows


def main():
    a = argparse.ArgumentParser()
    a.add_argument("plan")
    a.add_argument("--out", default="campaign.xlsx")
    args = a.parse_args()
    plan = json.load(open(args.plan, encoding="utf-8"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Тексты"
    head = [c for c, _ in COLUMNS]
    ws.append(head)
    fill = PatternFill("solid", fgColor="E2F4F1")
    for i, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True, size=10)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    gnum = 0
    for camp in plan["campaigns"]:
        for group in camp["groups"]:
            gnum += 1
            for row in rows_for_group(camp, group, gnum):
                ws.append([row.get(c, "") for c in head])

    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(args.out)
    print(f"групп {gnum}, строк {ws.max_row - 1}", file=sys.stderr)
    print(args.out)


if __name__ == "__main__":
    main()
